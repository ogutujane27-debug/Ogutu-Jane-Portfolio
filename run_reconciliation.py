import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Paths
base_dir = r"C:\Kubo"
csv_path = os.path.join(base_dir, "kubo_document_matching_report.csv")
excel_path = os.path.join(base_dir, "Kubo_Daily_Reconciliation_August_2026.xlsx")
latest_excel_path = os.path.join(base_dir, "Kubo_Today_Latest_Scans.xlsx")

print("1. Loading and cleaning data...")
if os.path.exists(csv_path):
    df = pd.read_csv(csv_path)
    # Clean any status inconsistencies
    if 'Status' in df.columns:
        df['Status'] = df['Status'].replace({'MISSING_IYES': 'MISSING_PDF'})
        df.to_csv(csv_path, index=False)
    
    print("2. Generating multi-tab Excel workbook...")
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df[df['Status'].str.contains('MATCH', na=False)].to_excel(writer, sheet_name='Matched Records', index=False)
        df[df['Status'].str.contains('MISSING', na=False)].to_excel(writer, sheet_name='Missing PDFs', index=False)
        df[df['Status'].str.contains('UNTRACKED', na=False)].to_excel(writer, sheet_name='Untracked PDFs', index=False)
        
    print("3. Exporting latest session snapshot...")
    df.tail(50).to_excel(latest_excel_path, index=False)

    print("4. Formatting workbook styles...")
    wb = openpyxl.load_workbook(excel_path)
    fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
    
    for ws in wb.worksheets:
        ws.freeze_panes = 'A2'
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=0)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    wb.save(excel_path)
    print("✅ Reconciliation processing complete! Opening Excel...")
    os.startfile(excel_path)
else:
    print(f"❌ Error: Could not find {csv_path}")